# src/report.py
"""Generación de reportes XLSX y HTML para informes DFIR."""

import os
from datetime import datetime
from collections import Counter

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#  XLSX
# ═══════════════════════════════════════════════════════════════

RISK_FILLS = {
    "CRITICAL": PatternFill(start_color="C0392B", fill_type="solid"),
    "HIGH": PatternFill(start_color="E67E22", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="F39C12", fill_type="solid"),
    "LOW": PatternFill(start_color="27AE60", fill_type="solid"),
    "SAFE": PatternFill(start_color="2980B9", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)


def generate_xlsx(results_list, case_name, output_folder):
    """Genera XLSX con pestañas: Resumen, Infraestructura, Red, WHOIS."""
    rows_summary, rows_infra, rows_net, rows_whois = [], [], [], []

    for r in results_list:
        ip = r["ip"]
        cip = r.get("criminal_ip", {})
        country = r.get("country_name", "N/A")

        # Merge TOR
        is_tor = r.get("is_tor_node") is True or cip.get("is_tor") is True

        rows_summary.append({
            "IP": ip,
            "Riesgo": r.get("risk_level", "N/A"),
            "Score": r.get("malicious_score", 0),
            "AbuseIPDB": r.get("ipabuse_score", "N/A"),
            "VirusTotal": r.get("vt_malicious", "N/A"),
            "CIP Inbound": cip.get("inbound_score", "N/A"),
            "CIP Outbound": cip.get("outbound_score", "N/A"),
            "TOR": is_tor,
            "País": country,
        })

        rows_infra.append({
            "IP": ip,
            "VPN": cip.get("is_vpn", "N/A"),
            "Nombre VPN": cip.get("vpn_name", "N/A"),
            "TOR": cip.get("is_tor", "N/A"),
            "Proxy": cip.get("is_proxy", "N/A"),
            "Hosting": cip.get("is_hosting", "N/A"),
            "Cloud": cip.get("is_cloud", "N/A"),
            "Scanner": cip.get("is_scanner", "N/A"),
            "Darkweb": cip.get("is_darkweb", "N/A"),
            "VPN Anónima": cip.get("is_anonymous_vpn", "N/A"),
            "Categorías": cip.get("categories", "N/A"),
        })

        rows_net.append({
            "IP": ip,
            "Puertos Abiertos": cip.get("open_ports_count", "N/A"),
            "Puertos Principales": cip.get("top_ports", "N/A"),
            "Vulnerabilidades": cip.get("vuln_count", "N/A"),
            "CVEs Principales": cip.get("top_cves", "N/A"),
            "Dominios": cip.get("domains_count", "N/A"),
            "Honeypot": cip.get("honeypot_count", "N/A"),
            "Alertas IDS": cip.get("ids_count", "N/A"),
        })

        rows_whois.append({
            "IP": ip,
            "País": country,
            "Código País": cip.get("country", "N/A"),
            "Ciudad": cip.get("city", "N/A"),
            "ASN": cip.get("asn", "N/A"),
            "Organización": cip.get("org", "N/A"),
        })

    path = os.path.join(output_folder, f"{case_name}_report.xlsx")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, rows in [
            ("Resumen", rows_summary),
            ("Infraestructura", rows_infra),
            ("Red", rows_net),
            ("WHOIS", rows_whois),
        ]:
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=name, index=False)

        # Formatear cada hoja
        wb = writer.book
        for ws in wb.worksheets:
            _format_sheet(ws)

    return path


def _format_sheet(ws):
    """Aplica formato profesional a una hoja."""
    # Headers
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-ajustar columnas y colorear riesgo
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        header_val = ws.cell(1, col_idx).value or ""

        for row_idx, cell in enumerate(col, 1):
            if row_idx > 1:
                cell.border = THIN_BORDER
                # Colorear columna "Riesgo"
                if header_val == "Riesgo" and cell.value in RISK_FILLS:
                    cell.fill = RISK_FILLS[cell.value]
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
                # Resaltar True en columnas booleanas
                if cell.value is True:
                    cell.font = Font(color="C0392B", bold=True)
                elif cell.value is False:
                    cell.font = Font(color="95A5A6")

            val_len = len(str(cell.value or ""))
            if val_len > max_len:
                max_len = val_len

        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
#  HTML
# ═══════════════════════════════════════════════════════════════

def generate_html(results_list, case_name, output_folder):
    """Genera HTML con gráficas Chart.js y tablas copiables para informes DFIR."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = len(results_list)

    # Conteos
    risk_counts = Counter(r.get("risk_level", "N/A") for r in results_list)
    country_counts = Counter(r.get("country_name", "N/A") for r in results_list)
    top_countries = country_counts.most_common(10)

    # Hallazgos
    cip_results = [r for r in results_list if isinstance(r.get("criminal_ip"), dict)]
    vpn_count = sum(1 for r in cip_results if r["criminal_ip"].get("is_vpn") is True)
    tor_count = sum(1 for r in results_list if r.get("is_tor_node") is True or (isinstance(r.get("criminal_ip"), dict) and r["criminal_ip"].get("is_tor") is True))
    proxy_count = sum(1 for r in cip_results if r["criminal_ip"].get("is_proxy") is True)
    scanner_count = sum(1 for r in cip_results if r["criminal_ip"].get("is_scanner") is True)
    darkweb_count = sum(1 for r in cip_results if r["criminal_ip"].get("is_darkweb") is True)
    hosting_count = sum(1 for r in cip_results if r["criminal_ip"].get("is_hosting") is True)
    vuln_ips = sum(1 for r in cip_results if isinstance(r["criminal_ip"].get("vuln_count"), int) and r["criminal_ip"]["vuln_count"] > 0)

    # Ordenar por score desc
    sorted_results = sorted(results_list, key=lambda x: x.get("malicious_score", 0), reverse=True)

    # ── Tablas HTML ──
    summary_rows = ""
    for r in sorted_results:
        cip = r.get("criminal_ip", {})
        risk = r.get("risk_level", "N/A")
        is_tor = r.get("is_tor_node") is True or cip.get("is_tor") is True
        summary_rows += f"""<tr>
            <td>{r['ip']}</td>
            <td><span class="badge {risk}">{risk}</span></td>
            <td>{r.get('malicious_score', 0)}</td>
            <td>{r.get('ipabuse_score', 'N/A')}</td>
            <td>{r.get('vt_malicious', 'N/A')}</td>
            <td>{cip.get('inbound_score', 'N/A')}</td>
            <td>{cip.get('outbound_score', 'N/A')}</td>
            <td class="{'flag-on' if is_tor else ''}">{is_tor}</td>
            <td>{r.get('country_name', 'N/A')}</td>
        </tr>"""

    infra_rows = ""
    for r in sorted_results:
        cip = r.get("criminal_ip", {})
        if not isinstance(cip, dict):
            continue

        def _flag(val):
            if val is True:
                return '<span class="flag-on">TRUE</span>'
            if val is False:
                return '<span class="flag-off">false</span>'
            return str(val)

        infra_rows += f"""<tr>
            <td>{r['ip']}</td>
            <td>{_flag(cip.get('is_vpn'))}</td>
            <td>{cip.get('vpn_name', 'N/A')}</td>
            <td>{_flag(cip.get('is_tor'))}</td>
            <td>{_flag(cip.get('is_proxy'))}</td>
            <td>{_flag(cip.get('is_hosting'))}</td>
            <td>{_flag(cip.get('is_cloud'))}</td>
            <td>{_flag(cip.get('is_scanner'))}</td>
            <td>{_flag(cip.get('is_darkweb'))}</td>
        </tr>"""

    net_rows = ""
    for r in sorted_results:
        cip = r.get("criminal_ip", {})
        if not isinstance(cip, dict):
            continue
        net_rows += f"""<tr>
            <td>{r['ip']}</td>
            <td>{cip.get('open_ports_count', 'N/A')}</td>
            <td class="small">{cip.get('top_ports', 'N/A')}</td>
            <td>{cip.get('vuln_count', 'N/A')}</td>
            <td class="small">{cip.get('top_cves', 'N/A')}</td>
            <td>{cip.get('domains_count', 'N/A')}</td>
            <td>{cip.get('honeypot_count', 'N/A')}</td>
            <td>{cip.get('ids_count', 'N/A')}</td>
        </tr>"""

    # ── Risk chart data ──
    risk_labels = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "SAFE"]
    risk_values = [risk_counts.get(r, 0) for r in risk_labels]
    risk_colors = ["#c0392b", "#e67e22", "#f39c12", "#27ae60", "#2980b9"]

    country_labels = [c[0] for c in top_countries]
    country_values = [c[1] for c in top_countries]

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>LogIPCore - {_esc(case_name)}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f0f2f5;color:#333;line-height:1.5}}
.wrap{{max-width:1280px;margin:0 auto;padding:24px}}
.header{{background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;padding:28px 32px;border-radius:10px;margin-bottom:24px}}
.header h1{{font-size:24px;font-weight:600}} .header .meta{{color:#94a3b8;font-size:13px;margin-top:4px}}
.cards{{display:flex;gap:12px;margin:20px 0;flex-wrap:wrap}}
.card{{flex:1;min-width:90px;padding:16px 12px;border-radius:8px;text-align:center;color:#fff}}
.card .num{{font-size:28px;font-weight:700}} .card .lbl{{font-size:11px;text-transform:uppercase;opacity:.85}}
.c-critical{{background:#c0392b}} .c-high{{background:#e67e22}} .c-medium{{background:#f39c12}}
.c-low{{background:#27ae60}} .c-safe{{background:#2980b9}} .c-total{{background:#34495e}}
.section{{background:#fff;border-radius:8px;padding:24px;margin:20px 0;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.section h2{{font-size:17px;color:#1a1a2e;border-bottom:2px solid #e2e8f0;padding-bottom:8px;margin-bottom:16px}}
.charts{{display:flex;gap:24px;flex-wrap:wrap}} .chart-box{{flex:1;min-width:280px;max-width:500px}}
.findings{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin:12px 0}}
.find{{padding:14px;background:#f8fafc;border-radius:6px;border-left:4px solid #3498db}}
.find.warn{{border-left-color:#e74c3c}} .find .v{{font-size:22px;font-weight:700;color:#1a1a2e}}
.find .d{{font-size:11px;color:#64748b}}
table{{width:100%;border-collapse:collapse;font-size:12px;margin-top:8px}}
thead th{{background:#1a1a2e;color:#fff;padding:9px 10px;text-align:left;font-weight:600;white-space:nowrap}}
tbody td{{padding:7px 10px;border-bottom:1px solid #f1f5f9}} tbody tr:hover{{background:#f8fafc}}
td.small{{font-size:11px;max-width:260px;word-break:break-all}}
.badge{{padding:2px 10px;border-radius:10px;font-size:11px;font-weight:700;color:#fff;white-space:nowrap}}
.CRITICAL{{background:#c0392b}} .HIGH{{background:#e67e22}} .MEDIUM{{background:#f39c12}}
.LOW{{background:#27ae60}} .SAFE{{background:#2980b9}}
.flag-on{{color:#c0392b;font-weight:700}} .flag-off{{color:#cbd5e1}}
.footer{{text-align:center;color:#94a3b8;font-size:11px;padding:20px;margin-top:16px}}
@media print{{body{{background:#fff}} .section{{box-shadow:none;border:1px solid #e2e8f0}} .wrap{{padding:0}}}}
</style>
</head>
<body>
<div class="wrap">

<div class="header">
  <h1>LogIPCore &mdash; Informe de Análisis</h1>
  <div class="meta">Caso: <strong>{_esc(case_name)}</strong> &nbsp;|&nbsp; {now} &nbsp;|&nbsp; {total} IPs analizadas</div>
</div>

<!-- Cards -->
<div class="cards">
  <div class="card c-total"><div class="num">{total}</div><div class="lbl">Total</div></div>
  <div class="card c-critical"><div class="num">{risk_counts.get('CRITICAL',0)}</div><div class="lbl">Critical</div></div>
  <div class="card c-high"><div class="num">{risk_counts.get('HIGH',0)}</div><div class="lbl">High</div></div>
  <div class="card c-medium"><div class="num">{risk_counts.get('MEDIUM',0)}</div><div class="lbl">Medium</div></div>
  <div class="card c-low"><div class="num">{risk_counts.get('LOW',0)}</div><div class="lbl">Low</div></div>
  <div class="card c-safe"><div class="num">{risk_counts.get('SAFE',0)}</div><div class="lbl">Safe</div></div>
</div>

<!-- Charts -->
<div class="section">
  <h2>Distribución</h2>
  <div class="charts">
    <div class="chart-box"><canvas id="riskChart"></canvas></div>
    <div class="chart-box"><canvas id="countryChart"></canvas></div>
  </div>
</div>

<!-- Findings -->
<div class="section">
  <h2>Hallazgos Principales</h2>
  <div class="findings">
    <div class="find{'warn' if vpn_count else ''}"><div class="v">{vpn_count}</div><div class="d">IPs con VPN</div></div>
    <div class="find{'warn' if tor_count else ''}"><div class="v">{tor_count}</div><div class="d">Nodos TOR</div></div>
    <div class="find{'warn' if proxy_count else ''}"><div class="v">{proxy_count}</div><div class="d">Proxies</div></div>
    <div class="find{'warn' if scanner_count else ''}"><div class="v">{scanner_count}</div><div class="d">Scanners</div></div>
    <div class="find{'warn' if darkweb_count else ''}"><div class="v">{darkweb_count}</div><div class="d">Darkweb</div></div>
    <div class="find"><div class="v">{hosting_count}</div><div class="d">Hosting/Cloud</div></div>
    <div class="find{'warn' if vuln_ips else ''}"><div class="v">{vuln_ips}</div><div class="d">IPs con vulns</div></div>
  </div>
</div>

<!-- Resumen -->
<div class="section">
  <h2>Resumen de Análisis</h2>
  <table>
    <thead><tr><th>IP</th><th>Riesgo</th><th>Score</th><th>AbuseIPDB</th><th>VirusTotal</th><th>CIP Inbound</th><th>CIP Outbound</th><th>TOR</th><th>País</th></tr></thead>
    <tbody>{summary_rows}</tbody>
  </table>
</div>

<!-- Infraestructura -->
<div class="section">
  <h2>Infraestructura</h2>
  <table>
    <thead><tr><th>IP</th><th>VPN</th><th>VPN Name</th><th>TOR</th><th>Proxy</th><th>Hosting</th><th>Cloud</th><th>Scanner</th><th>Darkweb</th></tr></thead>
    <tbody>{infra_rows}</tbody>
  </table>
</div>

<!-- Red -->
<div class="section">
  <h2>Red y Vulnerabilidades</h2>
  <table>
    <thead><tr><th>IP</th><th>Puertos</th><th>Top Puertos</th><th>Vulns</th><th>Top CVEs</th><th>Dominios</th><th>Honeypot</th><th>IDS</th></tr></thead>
    <tbody>{net_rows}</tbody>
  </table>
</div>

<div class="footer">Generado por LogIPCore &mdash; {now}</div>
</div>

<script>
const riskCtx = document.getElementById('riskChart');
if(riskCtx){{ new Chart(riskCtx, {{
  type:'doughnut',
  data:{{ labels:{risk_labels}, datasets:[{{ data:{risk_values}, backgroundColor:{risk_colors} }}] }},
  options:{{ responsive:true, plugins:{{ legend:{{position:'bottom'}}, title:{{display:true,text:'Distribución de Riesgo'}} }} }}
}}); }}

const countryCtx = document.getElementById('countryChart');
if(countryCtx){{ new Chart(countryCtx, {{
  type:'bar',
  data:{{ labels:{_js_list(country_labels)}, datasets:[{{ label:'IPs', data:{country_values}, backgroundColor:'#3498db' }}] }},
  options:{{ indexAxis:'y', responsive:true, plugins:{{ legend:{{display:false}}, title:{{display:true,text:'Top Países'}} }} }}
}}); }}
</script>
</body>
</html>"""

    path = os.path.join(output_folder, f"{case_name}_report.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path


def _esc(text):
    """Escape HTML."""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _js_list(items):
    """Convierte lista Python a literal JS string array."""
    escaped = [f'"{_esc(str(i))}"' for i in items]
    return "[" + ",".join(escaped) + "]"
